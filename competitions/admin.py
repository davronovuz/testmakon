"""
TestMakon.uz - Competitions Admin
Professional admin panel for competitions management
"""

from django.contrib import admin
from django.utils.html import format_html
from django.utils import timezone
from django.db.models import Count, Avg
from django.urls import reverse
from django.utils.safestring import mark_safe
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify
import uuid as uuid_lib

from .models import (
    Competition, CompetitionParticipant, CompetitionPayment,
    Certificate, Battle, BattleInvitation, MatchmakingQueue,
    DailyChallenge, DailyChallengeParticipant,
    WeeklyLeague, WeeklyLeagueParticipant,
    CompetitionQuestion
)


# ============================================================
# INLINE ADMINS
# ============================================================

class CompetitionParticipantInline(admin.TabularInline):
    model = CompetitionParticipant
    extra = 0
    readonly_fields = ['user', 'rank', 'score', 'percentage', 'correct_answers',
                       'wrong_answers', 'time_spent', 'xp_earned', 'status', 'joined_at']
    fields = ['user', 'status', 'rank', 'score', 'correct_answers', 'time_spent', 'xp_earned']
    can_delete = False
    max_num = 0

    def has_add_permission(self, request, obj=None):
        return False


class CertificateInline(admin.TabularInline):
    model = Certificate
    extra = 0
    readonly_fields = ['user', 'certificate_type', 'rank', 'score', 'verification_code', 'issued_at']
    fields = ['user', 'certificate_type', 'rank', 'verification_code']
    can_delete = False
    max_num = 0

    def has_add_permission(self, request, obj=None):
        return False


class DailyChallengeParticipantInline(admin.TabularInline):
    model = DailyChallengeParticipant
    extra = 0
    readonly_fields = ['user', 'rank', 'score', 'correct_answers', 'time_spent', 'xp_earned', 'completed_at']
    fields = ['user', 'rank', 'score', 'correct_answers', 'time_spent', 'xp_earned']
    can_delete = False
    max_num = 0

    def has_add_permission(self, request, obj=None):
        return False


class WeeklyLeagueParticipantInline(admin.TabularInline):
    model = WeeklyLeagueParticipant
    extra = 0
    readonly_fields = ['user', 'rank', 'xp_earned', 'tests_completed', 'is_promoted', 'is_demoted']
    fields = ['user', 'rank', 'xp_earned', 'tests_completed', 'is_promoted', 'is_demoted']
    can_delete = False
    max_num = 0

    def has_add_permission(self, request, obj=None):
        return False


class CompetitionQuestionInline(admin.TabularInline):
    model = CompetitionQuestion
    extra = 1
    autocomplete_fields = ['question']
    fields = ('order', 'question', 'question_preview')
    readonly_fields = ('question_preview',)
    ordering = ['order']

    def question_preview(self, obj):
        if obj.question_id:
            q = obj.question
            correct = q.answers.filter(is_correct=True).first()
            return format_html(
                '<small><b>Fan:</b> {} | <b>Qiyinlik:</b> {} | <b>Javob:</b> {}</small>',
                q.subject.name if q.subject else '-',
                q.get_difficulty_display(),
                correct.text[:40] if correct else '-'
            )
        return '-'
    question_preview.short_description = "Ma'lumot"


# ============================================================
# COMPETITION ADMIN
# ============================================================

@admin.register(Competition)
class CompetitionAdmin(admin.ModelAdmin):
    list_display = [
        'title', 'competition_type_badge', 'status_badge', 'entry_type_badge',
        'questions_count_display', 'participants_display', 'prize_display', 'start_time', 'is_active'
    ]
    list_filter = [
        'status', 'competition_type', 'entry_type', 'test_format', 'question_source',
        'is_active', 'is_featured', 'certificate_enabled', 'subject'
    ]
    change_form_template = 'admin/competitions/competition/change_form.html'
    search_fields = ['title', 'slug', 'description']
    prepopulated_fields = {'slug': ('title',)}
    date_hierarchy = 'start_time'
    ordering = ['-start_time']

    readonly_fields = [
        'uuid', 'participants_count', 'completed_count', 'average_score',
        'created_at', 'updated_at', 'status_info'
    ]

    fieldsets = (
        ('Asosiy ma\'lumotlar', {
            'fields': ('title', 'slug', 'icon', 'short_description', 'description', 'rules', 'banner')
        }),
        ('Tur va holat', {
            'fields': ('competition_type', 'status', 'entry_type', 'entry_fee', 'test_format')
        }),
        ('Fan va test', {
            'fields': ('subject', 'subjects', 'test', 'question_source',
                       'questions_per_subject', 'total_questions',
                       'difficulty_distribution'),
        }),
        ('Vaqt sozlamalari', {
            'fields': ('registration_start', 'registration_end', 'start_time', 'end_time', 'duration_minutes')
        }),
        ('Qatnashish shartlari', {
            'fields': ('max_participants', 'min_participants', 'min_level', 'min_rating'),
            'classes': ('collapse',)
        }),
        ('Sovrinlar', {
            'fields': ('prize_pool', 'prizes', 'xp_reward_first', 'xp_reward_second', 'xp_reward_third',
                       'xp_participation', 'xp_per_correct')
        }),
        ('Xususiyatlar', {
            'fields': ('show_live_leaderboard', 'show_answers_after', 'allow_review', 'certificate_enabled',
                       'anti_cheat_enabled'),
            'classes': ('collapse',)
        }),
        ('Homiylar', {
            'fields': ('sponsors',),
            'classes': ('collapse',)
        }),
        ('Sozlamalar', {
            'fields': ('is_active', 'is_featured', 'created_by')
        }),
        ('Statistika', {
            'fields': ('uuid', 'participants_count', 'completed_count', 'average_score', 'status_info', 'created_at',
                       'updated_at'),
            'classes': ('collapse',)
        }),
    )

    filter_horizontal = ['subjects']
    inlines = [CompetitionQuestionInline, CompetitionParticipantInline, CertificateInline]

    actions = ['make_active', 'make_finished', 'calculate_ranks', 'generate_certificates']

    def questions_count_display(self, obj):
        count = obj.competition_questions.count()
        if count > 0:
            return format_html(
                '<span style="color:#10B981; font-weight:600;">{} ta savol</span>', count
            )
        if obj.question_source == 'auto':
            return format_html(
                '<span style="color:#6B7280;">Avtomatik ({})</span>', obj.total_questions
            )
        return format_html('<span style="color:#EF4444;">Savol yo\'q!</span>')
    questions_count_display.short_description = 'Savollar'

    def get_urls(self):
        from django.urls import path
        custom_urls = [
            path(
                '<int:competition_id>/import-questions/',
                self.admin_site.admin_view(self.import_questions_view),
                name='competition-import-questions',
            ),
        ]
        return custom_urls + super().get_urls()

    def import_questions_view(self, request, competition_id):
        """Musobaqaga CSV/Excel dan savollar import qilish"""
        from tests_app.models import Subject, Topic, Question, Answer

        competition = get_object_or_404(Competition, pk=competition_id)

        if request.method == 'POST' and request.FILES.get('file'):
            file = request.FILES['file']
            import openpyxl
            import csv
            import io

            rows = []
            if file.name.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, row)))
                wb.close()
            else:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                rows = list(reader)

            created_count = 0
            for row in rows:
                savol_text = str(row.get('Savol', '')).strip()
                if not savol_text:
                    continue

                sub_name = str(row.get('Fan', 'Umumiy Fan')).strip()
                sub_slug = slugify(sub_name) or str(uuid_lib.uuid4())[:8]
                subject_obj, _ = Subject.objects.get_or_create(
                    name=sub_name,
                    defaults={'slug': sub_slug, 'is_active': True}
                )

                top_name = str(row.get('Mavzu', 'Umumiy Mavzu')).strip()
                top_slug = slugify(top_name) or str(uuid_lib.uuid4())[:8]
                topic_obj, _ = Topic.objects.get_or_create(
                    name=top_name, subject=subject_obj,
                    defaults={'slug': top_slug, 'is_active': True}
                )

                difficulty = str(row.get('Qiyinlik', 'medium')).strip().lower()
                if difficulty not in ('easy', 'medium', 'hard', 'expert'):
                    difficulty = 'medium'

                question = Question.objects.create(
                    text=savol_text,
                    subject=subject_obj,
                    topic=topic_obj,
                    difficulty=difficulty,
                    question_type='single',
                )

                correct_key = str(row.get('Togri_javob', '')).strip().upper()
                for idx, key in enumerate(['A', 'B', 'C', 'D']):
                    text = row.get(key)
                    if text and str(text).strip():
                        Answer.objects.create(
                            question=question,
                            text=str(text).strip(),
                            is_correct=(key == correct_key),
                            order=idx
                        )

                CompetitionQuestion.objects.create(
                    competition=competition,
                    question=question,
                    order=created_count
                )
                created_count += 1

            if competition.question_source == 'auto' and created_count > 0:
                competition.question_source = 'manual'
                competition.total_questions = created_count
                competition.save(update_fields=['question_source', 'total_questions'])

            self.message_user(request, f'{created_count} ta savol muvaffaqiyatli import qilindi!')
            return redirect(reverse('admin:competitions_competition_change', args=[competition_id]))

        context = {
            **self.admin_site.each_context(request),
            'competition': competition,
            'title': f'Savollarni import qilish: {competition.title}',
            'opts': self.model._meta,
            'existing_count': competition.competition_questions.count(),
        }
        return render(request, 'admin/competitions/import_questions.html', context)

    def competition_type_badge(self, obj):
        colors = {
            'daily': '#10B981',
            'weekly': '#3B82F6',
            'monthly': '#8B5CF6',
            'special': '#F59E0B',
            'olympiad': '#EF4444',
            'tournament': '#EC4899',
        }
        color = colors.get(obj.competition_type, '#6B7280')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; border-radius:12px; font-size:11px; font-weight:600;">{}</span>',
            color, obj.get_competition_type_display()
        )

    competition_type_badge.short_description = 'Turi'

    def status_badge(self, obj):
        colors = {
            'draft': '#6B7280',
            'upcoming': '#3B82F6',
            'registration': '#8B5CF6',
            'active': '#10B981',
            'finished': '#6B7280',
            'cancelled': '#EF4444',
        }
        color = colors.get(obj.status, '#6B7280')
        return format_html(
            '<span style="background:{}; color:white; padding:3px 10px; border-radius:12px; font-size:11px; font-weight:600;">{}</span>',
            color, obj.get_status_display()
        )

    status_badge.short_description = 'Holat'

    def entry_type_badge(self, obj):
        if obj.entry_type == 'free':
            return format_html('<span style="color:#10B981; font-weight:600;">‚úì Bepul</span>')
        elif obj.entry_type == 'premium_only':
            return format_html('<span style="color:#F59E0B; font-weight:600;">üëë Premium</span>')
        else:
            return format_html('<span style="color:#EF4444; font-weight:600;">üí∞ {} so\'m</span>', obj.entry_fee)

    entry_type_badge.short_description = 'Kirish'

    def participants_display(self, obj):
        if obj.max_participants:
            return format_html(
                '<span style="font-weight:600;">{}</span> / {}',
                obj.participants_count, obj.max_participants
            )
        return format_html('<span style="font-weight:600;">{}</span>', obj.participants_count)

    participants_display.short_description = 'Qatnashchilar'

    def prize_display(self, obj):
        if obj.prize_pool > 0:
            return format_html(
                '<span style="color:#10B981; font-weight:600;">{:,} so\'m</span>',
                obj.prize_pool
            )
        return '-'

    prize_display.short_description = 'Sovrin'

    def status_info(self, obj):
        now = timezone.now()
        if obj.status == 'upcoming' and obj.start_time > now:
            delta = obj.start_time - now
            days = delta.days
            hours = delta.seconds // 3600
            return format_html('<span style="color:#3B82F6;">‚è≥ {} kun {} soat qoldi</span>', days, hours)
        elif obj.status == 'active' and obj.end_time > now:
            delta = obj.end_time - now
            hours = delta.seconds // 3600
            minutes = (delta.seconds % 3600) // 60
            return format_html('<span style="color:#10B981;">üü¢ {} soat {} daqiqa qoldi</span>', hours, minutes)
        elif obj.status == 'finished':
            return format_html('<span style="color:#6B7280;">‚úÖ Yakunlangan</span>')
        return '-'

    status_info.short_description = 'Holat ma\'lumoti'

    @admin.action(description='Faollashtirish')
    def make_active(self, request, queryset):
        queryset.update(status='active')
        self.message_user(request, f'{queryset.count()} ta musobaqa faollashtirildi.')

    @admin.action(description='Yakunlash')
    def make_finished(self, request, queryset):
        queryset.update(status='finished')
        self.message_user(request, f'{queryset.count()} ta musobaqa yakunlandi.')

    @admin.action(description='O\'rinlarni hisoblash')
    def calculate_ranks(self, request, queryset):
        for competition in queryset:
            participants = CompetitionParticipant.objects.filter(
                competition=competition,
                status='completed'
            ).order_by('-score', 'time_spent')

            for rank, participant in enumerate(participants, 1):
                participant.rank = rank

                # XP hisoblash
                if rank == 1:
                    participant.xp_earned = competition.xp_reward_first
                elif rank == 2:
                    participant.xp_earned = competition.xp_reward_second
                elif rank == 3:
                    participant.xp_earned = competition.xp_reward_third
                else:
                    participant.xp_earned = competition.xp_participation

                # Sovrin
                prize = competition.get_prize_for_rank(rank)
                if prize:
                    participant.prize_amount = prize.get('amount', 0)

                participant.save()

            # Statistikani yangilash
            competition.completed_count = participants.count()
            if participants.exists():
                competition.average_score = participants.aggregate(Avg('score'))['score__avg'] or 0
            competition.save()

        self.message_user(request, f'{queryset.count()} ta musobaqa uchun o\'rinlar hisoblandi.')

    @admin.action(description='Sertifikatlar yaratish')
    def generate_certificates(self, request, queryset):
        count = 0
        for competition in queryset.filter(certificate_enabled=True):
            participants = CompetitionParticipant.objects.filter(
                competition=competition,
                status='completed'
            ).select_related('user')

            for participant in participants:
                if hasattr(participant, 'certificate'):
                    continue

                # Sertifikat turini aniqlash
                if participant.rank == 1:
                    cert_type = 'winner'
                elif participant.rank and participant.rank <= 3:
                    cert_type = 'top3'
                elif participant.rank and participant.rank <= 10:
                    cert_type = 'top10'
                else:
                    cert_type = 'participation'

                Certificate.objects.create(
                    user=participant.user,
                    competition=competition,
                    participant=participant,
                    certificate_type=cert_type,
                    rank=participant.rank,
                    score=participant.score
                )
                count += 1

        self.message_user(request, f'{count} ta sertifikat yaratildi.')


# ============================================================
# COMPETITION PARTICIPANT ADMIN
# ============================================================

@admin.register(CompetitionParticipant)
class CompetitionParticipantAdmin(admin.ModelAdmin):
    list_display = [
        'user', 'competition', 'status_badge', 'rank_display',
        'score_display', 'correct_answers', 'time_display', 'xp_earned', 'joined_at'
    ]
    list_filter = ['status', 'competition', 'is_suspected']
    search_fields = ['user__first_name', 'user__last_name', 'user__phone', 'competition__title']
    ordering = ['-joined_at']
    readonly_fields = [
        'competition', 'user', 'score', 'percentage', 'correct_answers',
        'wrong_answers', 'skipped_answers', 'time_spent', 'rank', 'xp_earned',
        'prize_amount', 'started_at', 'completed_at', 'violations_count',
        'violations_log', 'answers_data', 'joined_at'
    ]

    def status_badge(self, obj):
        colors = {
            'registered': '#6B7280',
            'ready': '#3B82F6',
            'in_progress': '#F59E0B',
            'completed': '#10B981',
            'disqualified': '#EF4444',
        }
        color = colors.get(obj.status, '#6B7280')
        return format_html(
            '<span style="background:{}; color:white; padding:2px 8px; border-radius:10px; font-size:11px;">{}</span>',
            color, obj.get_status_display()
        )

    status_badge.short_description = 'Holat'

    def rank_display(self, obj):
        if obj.rank:
            if obj.rank == 1:
                return format_html('<span style="font-size:16px;">ü•á 1</span>')
            elif obj.rank == 2:
                return format_html('<span style="font-size:16px;">ü•à 2</span>')
            elif obj.rank == 3:
                return format_html('<span style="font-size:16px;">ü•â 3</span>')
            return f'#{obj.rank}'
        return '-'

    rank_display.short_description = 'O\'rin'

    def score_display(self, obj):
        return format_html(
            '<span style="font-weight:600;">{}</span> <span style="color:#6B7280;">({:.1f}%)</span>',
            obj.score, obj.percentage
        )

    score_display.short_description = 'Ball'

    def time_display(self, obj):
        if obj.time_spent:
            minutes = obj.time_spent // 60
            seconds = obj.time_spent % 60
            return f'{minutes}:{seconds:02d}'
        return '-'

    time_display.short_description = 'Vaqt'


# ============================================================
# PAYMENT ADMIN
# ============================================================

@admin.register(CompetitionPayment)
class CompetitionPaymentAdmin(admin.ModelAdmin):
    list_display = ['uuid', 'participant_user', 'participant_competition', 'amount_display', 'status_badge',
                    'payment_method', 'paid_at']
    list_filter = ['status', 'payment_method']
    search_fields = ['uuid', 'participant__user__phone', 'transaction_id']
    ordering = ['-created_at']
    readonly_fields = ['uuid', 'participant', 'amount', 'created_at']

    def participant_user(self, obj):
        return obj.participant.user

    participant_user.short_description = 'Foydalanuvchi'

    def participant_competition(self, obj):
        return obj.participant.competition.title

    participant_competition.short_description = 'Musobaqa'

    def amount_display(self, obj):
        return format_html('<span style="font-weight:600;">{:,} so\'m</span>', obj.amount)

    amount_display.short_description = 'Summa'

    def status_badge(self, obj):
        colors = {
            'pending': '#F59E0B',
            'completed': '#10B981',
            'failed': '#EF4444',
            'refunded': '#6B7280',
        }
        color = colors.get(obj.status, '#6B7280')
        return format_html(
            '<span style="background:{}; color:white; padding:2px 8px; border-radius:10px; font-size:11px;">{}</span>',
            color, obj.get_status_display()
        )

    status_badge.short_description = 'Holat'


# ============================================================
# CERTIFICATE ADMIN
# ============================================================

@admin.register(Certificate)
class CertificateAdmin(admin.ModelAdmin):
    list_display = ['verification_code', 'user', 'competition', 'type_badge', 'rank_display', 'score', 'issued_at']
    list_filter = ['certificate_type', 'competition', 'is_verified']
    search_fields = ['verification_code', 'user__first_name', 'user__last_name', 'user__phone']
    ordering = ['-issued_at']
    readonly_fields = ['uuid', 'verification_code', 'issued_at']

    def type_badge(self, obj):
        colors = {
            'winner': '#F59E0B',
            'top3': '#8B5CF6',
            'top10': '#3B82F6',
            'participation': '#6B7280',
        }
        icons = {
            'winner': 'üèÜ',
            'top3': 'ü•â',
            'top10': 'üéñÔ∏è',
            'participation': 'üìú',
        }
        color = colors.get(obj.certificate_type, '#6B7280')
        icon = icons.get(obj.certificate_type, 'üìú')
        return format_html(
            '<span style="background:{}; color:white; padding:2px 8px; border-radius:10px; font-size:11px;">{} {}</span>',
            color, icon, obj.get_certificate_type_display()
        )

    type_badge.short_description = 'Turi'

    def rank_display(self, obj):
        if obj.rank:
            if obj.rank == 1:
                return 'ü•á 1-o\'rin'
            elif obj.rank == 2:
                return 'ü•à 2-o\'rin'
            elif obj.rank == 3:
                return 'ü•â 3-o\'rin'
            return f'#{obj.rank}'
        return '-'

    rank_display.short_description = 'O\'rin'


# ============================================================
# BATTLE ADMIN
# ============================================================

@admin.register(Battle)
class BattleAdmin(admin.ModelAdmin):
    list_display = [
        'uuid_short', 'challenger', 'vs_display', 'opponent_display',
        'subject', 'status_badge', 'winner_display', 'created_at'
    ]
    list_filter = ['status', 'opponent_type', 'bot_difficulty', 'subject', 'is_dtm_format']
    search_fields = ['uuid', 'challenger__first_name', 'opponent__first_name']
    ordering = ['-created_at']
    readonly_fields = [
        'uuid', 'challenger_score', 'challenger_correct', 'challenger_time',
        'challenger_answers', 'challenger_completed', 'opponent_score',
        'opponent_correct', 'opponent_time', 'opponent_answers', 'opponent_completed',
        'winner', 'is_draw', 'winner_is_bot', 'xp_awarded', 'questions_data',
        'created_at', 'accepted_at', 'started_at', 'completed_at'
    ]

    fieldsets = (
        ('O\'yinchilar', {
            'fields': ('challenger', 'opponent', 'opponent_type', 'bot_difficulty')
        }),
        ('Sozlamalar', {
            'fields': ('subject', 'is_dtm_format', 'question_count', 'time_per_question', 'total_time')
        }),
        ('Holat', {
            'fields': ('status', 'expires_at')
        }),
        ('Chaqiruvchi natijasi', {
            'fields': ('challenger_score', 'challenger_correct', 'challenger_time', 'challenger_completed'),
            'classes': ('collapse',)
        }),
        ('Raqib natijasi', {
            'fields': ('opponent_score', 'opponent_correct', 'opponent_time', 'opponent_completed'),
            'classes': ('collapse',)
        }),
        ('Natija', {
            'fields': ('winner', 'is_draw', 'winner_is_bot', 'winner_xp', 'loser_xp', 'xp_awarded')
        }),
        ('Vaqtlar', {
            'fields': ('created_at', 'accepted_at', 'started_at', 'completed_at'),
            'classes': ('collapse',)
        }),
    )

    def uuid_short(self, obj):
        return str(obj.uuid)[:8]

    uuid_short.short_description = 'ID'

    def vs_display(self, obj):
        return format_html('<span style="color:#6B7280; font-weight:600;">VS</span>')

    vs_display.short_description = ''

    def opponent_display(self, obj):
        if obj.opponent_type == 'bot':
            colors = {'easy': '#10B981', 'medium': '#F59E0B', 'hard': '#EF4444', 'expert': '#8B5CF6'}
            color = colors.get(obj.bot_difficulty, '#6B7280')
            return format_html(
                '<span style="color:{};">ü§ñ Bot ({})</span>',
                color, obj.get_bot_difficulty_display()
            )
        return obj.opponent or '-'

    opponent_display.short_description = 'Raqib'

    def status_badge(self, obj):
        colors = {
            'pending': '#F59E0B',
            'searching': '#3B82F6',
            'accepted': '#8B5CF6',
            'in_progress': '#10B981',
            'completed': '#6B7280',
            'rejected': '#EF4444',
            'expired': '#6B7280',
            'cancelled': '#EF4444',
        }
        color = colors.get(obj.status, '#6B7280')
        return format_html(
            '<span style="background:{}; color:white; padding:2px 8px; border-radius:10px; font-size:11px;">{}</span>',
            color, obj.get_status_display()
        )

    status_badge.short_description = 'Holat'

    def winner_display(self, obj):
        if obj.is_draw:
            return format_html('<span style="color:#F59E0B;">ü§ù Durrang</span>')
        elif obj.winner_is_bot:
            return format_html('<span style="color:#EF4444;">ü§ñ Bot yutdi</span>')
        elif obj.winner:
            return format_html('<span style="color:#10B981;">üèÜ {}</span>', obj.winner)
        return '-'

    winner_display.short_description = 'G\'olib'


# ============================================================
# MATCHMAKING ADMIN
# ============================================================

@admin.register(MatchmakingQueue)
class MatchmakingQueueAdmin(admin.ModelAdmin):
    list_display = ['user', 'subject', 'user_rating', 'question_count', 'is_matched', 'joined_at', 'expires_at']
    list_filter = ['is_matched', 'subject']
    search_fields = ['user__first_name', 'user__phone']
    ordering = ['-joined_at']


# ============================================================
# DAILY CHALLENGE ADMIN
# ============================================================

@admin.register(DailyChallenge)
class DailyChallengeAdmin(admin.ModelAdmin):
    list_display = ['date', 'subject', 'question_count', 'participants_count', 'average_score_display', 'xp_reward',
                    'is_active']
    list_filter = ['is_active', 'subject']
    search_fields = ['date']
    ordering = ['-date']
    filter_horizontal = ['questions']
    inlines = [DailyChallengeParticipantInline]

    def average_score_display(self, obj):
        return f'{obj.average_score:.1f}%'

    average_score_display.short_description = 'O\'rtacha ball'


@admin.register(DailyChallengeParticipant)
class DailyChallengeParticipantAdmin(admin.ModelAdmin):
    list_display = ['user', 'challenge', 'rank', 'score', 'correct_answers', 'time_display', 'xp_earned',
                    'completed_at']
    list_filter = ['challenge']
    search_fields = ['user__first_name', 'user__phone']
    ordering = ['-completed_at']

    def time_display(self, obj):
        if obj.time_spent:
            minutes = obj.time_spent // 60
            seconds = obj.time_spent % 60
            return f'{minutes}:{seconds:02d}'
        return '-'

    time_display.short_description = 'Vaqt'


# ============================================================
# WEEKLY LEAGUE ADMIN
# ============================================================

@admin.register(WeeklyLeague)
class WeeklyLeagueAdmin(admin.ModelAdmin):
    list_display = ['week_start', 'week_end', 'tier_badge', 'participants_count', 'xp_reward_first', 'is_active',
                    'is_processed']
    list_filter = ['tier', 'is_active', 'is_processed']
    ordering = ['-week_start']
    inlines = [WeeklyLeagueParticipantInline]

    actions = ['process_league']

    def tier_badge(self, obj):
        colors = {
            'bronze': '#CD7F32',
            'silver': '#C0C0C0',
            'gold': '#FFD700',
            'platinum': '#E5E4E2',
            'diamond': '#B9F2FF',
            'master': '#FF6B6B',
            'grandmaster': '#9B59B6',
        }
        color = colors.get(obj.tier, '#6B7280')
        return format_html(
            '<span style="background:{}; color:#1F2937; padding:3px 10px; border-radius:12px; font-size:11px; font-weight:600;">{}</span>',
            color, obj.get_tier_display()
        )

    tier_badge.short_description = 'Daraja'

    def participants_count(self, obj):
        return obj.participants.count()

    participants_count.short_description = 'Qatnashchilar'

    @admin.action(description='Ligani hisoblash')
    def process_league(self, request, queryset):
        for league in queryset.filter(is_processed=False):
            participants = WeeklyLeagueParticipant.objects.filter(
                league=league
            ).order_by('-xp_earned')

            for rank, participant in enumerate(participants, 1):
                participant.rank = rank

                # Ko'tarilish/tushish
                if rank <= league.promotion_count:
                    participant.is_promoted = True
                elif rank > participants.count() - league.demotion_count:
                    participant.is_demoted = True

                participant.save()

            league.is_processed = True
            league.save()

        self.message_user(request, f'{queryset.count()} ta liga hisoblandi.')


@admin.register(WeeklyLeagueParticipant)
class WeeklyLeagueParticipantAdmin(admin.ModelAdmin):
    list_display = ['user', 'league', 'rank', 'xp_earned', 'tests_completed', 'status_display']
    list_filter = ['league', 'is_promoted', 'is_demoted']
    search_fields = ['user__first_name', 'user__phone']
    ordering = ['rank']

    def status_display(self, obj):
        if obj.is_promoted:
            return format_html('<span style="color:#10B981;">‚¨ÜÔ∏è Ko\'tarildi</span>')
        elif obj.is_demoted:
            return format_html('<span style="color:#EF4444;">‚¨áÔ∏è Tushdi</span>')
        return format_html('<span style="color:#6B7280;">‚û°Ô∏è Qoldi</span>')

    status_display.short_description = 'Holat'


# ============================================================
# BATTLE INVITATION ADMIN
# ============================================================

@admin.register(BattleInvitation)
class BattleInvitationAdmin(admin.ModelAdmin):
    list_display = ['battle', 'sent_at', 'seen_at', 'responded_at', 'notification_sent']
    list_filter = ['notification_sent']
    ordering = ['-sent_at']